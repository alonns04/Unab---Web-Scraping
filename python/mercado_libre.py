import requests
from bs4 import BeautifulSoup
import openpyxl
import re
import pandas as pd

class MercadoLibre():

    def __init__(self, product_name: str):
        self.product_name = product_name
        self.array_products = [] # Lista con los productos y sus características. precio, nombre, etc.
        self.link_product = self.get_link_product()
        self.requests = requests.get(self.get_link_product())
        self.html_content = self.requests.content
        self.soup_parsed = BeautifulSoup(self.html_content, 'html.parser') # Contenido Parseado
        self.all_pages()
        self.divs = self.soup_parsed.find_all('div', class_='ui-search-result__content-wrapper') # Filtra según etiqueta y clase
        self.create_filter_list()
        self.create_excel()

    def all_pages(self):
        next_link = self.soup_parsed.find('a', {'title': 'Siguiente'})
        index = 0
        if next_link:
            next_page_url = next_link.get('href')
            while next_page_url:
                    if index < 10:
                        new_request = requests.get(next_page_url)
                        new_html_content = new_request.content
                        new_soup_parsed = BeautifulSoup(new_html_content, 'html.parser')
                        self.soup_parsed.append(new_soup_parsed)
                        next_link = new_soup_parsed.find('a', {'title': 'Siguiente'})
                        index += 1
                    else:
                        break

    def get_product_name(self):
        return self.product_name
    
    def get_link_product(self):
        def meli_string1(string): # Reemplaza los espacios por _
            string = string.strip().lower()
            return re.sub(r'\s+', '-', string)
        def meli_string2(string): # Reemplaza los espacios por "20%"
            return string.replace(" ", "%20")
        
        return 'https://listado.mercadolibre.com.ar/{}#D[A:{}]'.format(meli_string1(self.get_product_name()), meli_string2(self.get_product_name()))
    
    def create_filter_list(self):
        for i in self.divs: # Itera en los divs de la página y extra la información
            producto = i.find("h2")
            moneda = i.find("span", class_="andes-money-amount__currency-symbol")
            precio = i.find("span", class_="andes-money-amount__fraction")
            link = i.find("a")["href"]
            if producto and precio:
                self.array_products.append({"nombre": producto.text.strip(), "precio": precio.text.strip(), "moneda": moneda.text.strip(), "link": link})
        # Filtrar elementos que comiencen con "https://click1." (se repiten)
        self.array_products = [producto for producto in self.array_products if not producto['link'].startswith('https://click1.')]
    
    def create_excel(self):
            workbook = openpyxl.Workbook() # Genera un excel
            active_sheet = workbook.active # Abre el excel        
            excel_path = '../excel/producto-{}-MELI.xlsx'.format(self.get_product_name().replace(" ","_").replace("-","_"))
            active_sheet.cell(row = 1, column = 1, value = "nombre")
            active_sheet.cell(row = 1, column = 2, value = "moneda")
            active_sheet.cell(row = 1, column = 3, value = "precio")
            active_sheet.cell(row = 1, column = 4, value = "link")
            for i in range(len(self.array_products)): # Itera y lo va metiendo en el excel
                active_sheet.cell(row = i + 2, column = 1, value = self.array_products[i]["nombre"])
                active_sheet.cell(row = i + 2, column = 2, value =  self.array_products[i]["moneda"])
                active_sheet.cell(row = i + 2, column = 3, value =  self.array_products[i]["precio"])
                active_sheet.cell(row = i + 2, column = 4, value =  self.array_products[i]["link"])
            workbook.save(excel_path)
    def __str__(self):
        return self.array_products