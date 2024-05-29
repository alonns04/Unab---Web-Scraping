import requests
from bs4 import BeautifulSoup
import openpyxl
import re

class MercadoLibre():

    def __init__(self, product_name: str):
        self.product_name = product_name
        self.array_products = [] # Lista con los productos y sus características. precio, nombre, etc.
        self.link_product = self.get_link_product()
        self.requests = requests.get(self.get_link_product())
        self.html_content = self.requests.content
        self.soup_parsed = BeautifulSoup(self.html_content, 'html.parser') # Contenido Parseado
        self.divs = self.soup_parsed.find_all('div', class_='ui-search-result__content-wrapper') # Filtra según etiqueta y clase

        self.workbook = openpyxl.Workbook() # Genera un excel
        self.active_sheet = self.workbook.active # Abre el excel        
        self.excel_path = '../excel/producto-{}.xlsx'.format(self.get_product_name().replace(" ","_").replace("-","_"))


        self.create_filter_list()
        self.create_excel()

    def get_product_name(self):
        return self.product_name
    
    def get_link_product(self):
        def meli_string1(string): # Reemplaza los espacios por _
            string = string.strip().lower()
            return re.sub(r'\s+', '-', string)
        def meli_string2(string): # Reemplaza los espacios por "20%"
            return string.replace(" ", "%20")
        
        return 'https://listado.mercadolibre.com.ar/{}#D[A:{}]'.format(meli_string1(self.get_product_name()), meli_string2(self.get_product_name()))
    
    def create_filter_list(self):
        for i in self.divs: # Itera en los divs de la página y extra la información
            producto = i.find("h2")
            moneda = i.find("span", class_="andes-money-amount__currency-symbol")
            precio = i.find("span", class_="andes-money-amount__fraction")
            link = i.find("a")["href"]
            if producto and precio:
                self.array_products.append({"nombre": producto.text.strip(), "precio": precio.text.strip(), "moneda": moneda.text.strip(), "link": link})
        # Filtrar elementos que comiencen con "https://click1." (se repiten)
        self.array_products = [producto for producto in self.array_products if not producto['link'].startswith('https://click1.')]

    def create_excel(self):
        self.active_sheet.cell(row = 1, column = 1, value = "nombre")
        self.active_sheet.cell(row = 1, column = 2, value = "moneda")
        self.active_sheet.cell(row = 1, column = 3, value = "precio")
        self.active_sheet.cell(row = 1, column = 4, value = "link")
        for i in range(len(self.array_products)): # Itera y lo va metiendo en el excel
            self.active_sheet.cell(row = i + 2, column = 1, value = self.array_products[i]["nombre"])
            self.active_sheet.cell(row = i + 2, column = 2, value =  self.array_products[i]["moneda"])
            self.active_sheet.cell(row = i + 2, column = 3, value =  self.array_products[i]["precio"])
            self.active_sheet.cell(row = i + 2, column = 4, value =  self.array_products[i]["link"])
        self.workbook.save(self.excel_path)

producto_1 = MercadoLibre(input("Qué buscamos? : "))
producto_2 = MercadoLibre(input("Qué buscamos? : "))
