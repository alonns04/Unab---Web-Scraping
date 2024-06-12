from bs4 import BeautifulSoup
import requests
import time
import datetime
import smtplib
import openpyxl



class Amazon():
    def __init__(self, product_name: str):
        self.product_name = product_name
        self.array_products = [] # Lista con los productos y sus características. precio, nombre, etc.
        self.link_product = self.generar_link_amazon()
        self.headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36 OPR/109.0.0.0", "Accept-Encoding":"gzip, deflate, br, zstd", "Accept":"text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7", "DNT":"1","Connection":"close", "Upgrade-Insecure-Requests":"1"}
        self.requests = requests.get(self.get_link_product(), headers=self.headers)
        self.html_content = BeautifulSoup(self.requests.content, "html.parser")
        self.all_pages()
        self.divs = self.html_content.find_all("div", class_="sg-col-20-of-24 s-result-item s-asin sg-col-0-of-12 sg-col-16-of-20 sg-col s-widget-spacing-small gsx-ies-anchor sg-col-12-of-16")
        self.create_filter_list()
        self.create_excel()


    def get_link_product(self):
        return self.link_product
        
    def get_product_name(self):
        return self.product_name
    def generar_link_amazon(self):
        texto_limpio = self.product_name.strip().replace(" ", "+")
        base_url = "https://www.amazon.com/s?k="
        url_completa = f"{base_url}{texto_limpio}&__mk_es_US=ÅMÅŽÕÑ&crid=&sprefix=&ref=nb_sb_noss"
        return url_completa
    
    
    def create_filter_list(self):
        for i in self.divs: # Itera en los divs de la página y extra la información
            producto = i.find("span", class_="a-size-medium a-color-base a-text-normal")
            link = "https://www.amazon.com/-/es" + i.find("a", class_="a-link-normal s-underline-text s-underline-link-text s-link-style a-text-normal").get("href")
            #moneda = i.find("span", class_="a-offscreen")
            #if moneda:
            #    moneda = moneda.text
            moneda = "US$"
            price = i.find("span", class_="a-price-whole")
            cents = i.find("span", class_="a-price-fraction")
            if price and price.text:
                price = price.text.split(".")[0] 
                if cents:
                    price = int(price) + int(cents.text) / 100
                self.array_products.append({
                    "nombre": producto.text,
                    "precio": price,
                    "moneda": moneda,
                    "link": link
                })
    
    def all_pages(self):
        next_link = self.html_content.find('a', {'class': 's-pagination-item s-pagination-next s-pagination-button s-pagination-separator'})
        index = 0
        if next_link:
            next_page_url = "https://www.amazon.com/-/es" + next_link.get('href')
            while next_page_url:
                if index < 10:
                    new_request = requests.get(next_page_url, headers=self.headers)
                    new_html_content = new_request.content
                    new_soup_parsed = BeautifulSoup(new_html_content, 'html.parser')
                    self.html_content.append(new_soup_parsed)
                    next_link = self.html_content.find('a', {'class': 's-pagination-item s-pagination-next s-pagination-button s-pagination-separator'})
                    index += 1
                else: 
                    break
    
    def create_excel(self):
            workbook = openpyxl.Workbook() # Genera un excel
            active_sheet = workbook.active # Abre el excel        
            excel_path = '../excel/producto-{}-AMZ.xlsx'.format(self.get_product_name().replace(" ","_").replace("-","_"))
            active_sheet.cell(row = 1, column = 1, value = "nombre")
            active_sheet.cell(row = 1, column = 2, value = "moneda")
            active_sheet.cell(row = 1, column = 3, value = "precio")
            active_sheet.cell(row = 1, column = 4, value = "link")
            for i in range(len(self.array_products)): # Itera y lo va metiendo en el excel
                active_sheet.cell(row = i + 2, column = 1, value = self.array_products[i]["nombre"])
                active_sheet.cell(row = i + 2, column = 2, value =  self.array_products[i]["moneda"])
                active_sheet.cell(row = i + 2, column = 3, value =  self.array_products[i]["precio"])
                active_sheet.cell(row = i + 2, column = 4, value =  self.array_products[i]["link"])
            workbook.save(excel_path)
    def __str__(self):
        return self.array_products