import openpyxl
import os

def create_excel(product):
        workbook = openpyxl.Workbook() # Genera un excel
        active_sheet = workbook.active # Abre el excel  
        relative_path = os.path.join('..', 'excel')
        absolute_path = os.path.abspath(relative_path) + product.excel_path
        active_sheet.cell(row = 1, column = 1, value = "nombre")
        active_sheet.cell(row = 1, column = 2, value = "moneda")
        active_sheet.cell(row = 1, column = 3, value = "precio")
        active_sheet.cell(row = 1, column = 4, value = "link")
        for i in range(len(product.array_products)): # Itera y lo va metiendo en el excel
            active_sheet.cell(row = i + 2, column = 1, value = product.array_products[i]["nombre"])
            active_sheet.cell(row = i + 2, column = 2, value =  product.array_products[i]["moneda"])
            active_sheet.cell(row = i + 2, column = 3, value =  product.array_products[i]["precio"])
            active_sheet.cell(row = i + 2, column = 4, value =  product.array_products[i]["link"])
        workbook.save(absolute_path)